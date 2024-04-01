import openpyxl as xl


from openpyxl.chart import BarChart, Reference
#openpyxl - >package
#chart -> module 
#BarChart...Reference --> classes



wb = xl.load_workbook("transactions.xlsx")     #xcell itself

sheet = wb["Sheet1"] #sheet



#19-20 not needed just to demonstrate how to get a cell

cell = sheet['a1'] #x then y
cell = sheet.cell (1,2) #lines 11 and 12 do the same thing #coordinate: row then col


print (cell.value) #prints what is in the cell



print (sheet.max_row) #tells max # of rows here it is 4

print()
print()
print()

for row in range (2,sheet.max_row+1): #for all the prices at the col 3 (rxcluding the first one as it is a header )
    cell = sheet.cell(row,3)
    #print (cell.value)
    corrected_price = cell.value * (0.9)
    corrected_cell = sheet.cell (row, 4)
    corrected_cell.value = corrected_price


values =  Reference (sheet, min_row=2, max_row = sheet.max_row, min_col = 4, max_col = 4) # sheet, min row, max row, min col, max col



chart = BarChart () #creates a new bar chart
chart.add_data(values) #acreates graph from data from Reference
sheet.add_chart (chart, 'E2')



wb.save ("transactions22.xlsx") #saves wb as a file (paramater makes it a diff one)

